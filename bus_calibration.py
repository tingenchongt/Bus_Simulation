"""
Read Data_Collection.xlsx and build UXsim inputs from all survey columns.

Location (where observation started):
  - Robinsons Dasmarinas (etc.) -> corridor context Robinsons -> Waltermart; rows drive sim 1-2 calibration.
  - Waltermart Dasmarinas (etc.) -> Waltermart -> Robinsons context; rows drive sim 3-4 calibration.

simulation.py calls load_session_stats(sheet_name, data_origin=...).
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Literal

import openpyxl

DATA_PATH = Path(__file__).resolve().parent / "Data_Collection.xlsx"
SESSION_SHEETS = ("Morning Session", "Lunch", "Afternoon Session")

DataOrigin = Literal["robinsons", "waltermart"]

SIM_CLEARANCE_S = 10800.0
MIN_DEMAND_SPAN_S = 1800.0
MAX_DEMAND_WINDOW_CAP_S = 5 * 3600.0


def _time_to_seconds(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, datetime):
        return _time_to_seconds(val.time())
    s = str(val).strip()
    m = re.match(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        h, mn, sec = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        return h * 3600 + mn * 60 + sec
    return None


def _normalize_location(loc: str) -> str:
    t = (loc or "").strip().lower()
    if "robinson" in t:
        return "robinsons"
    if "walter" in t:
        return "waltermart"
    return "other"


def parse_dwell_seconds(raw):
    """Parse dwell cells like '30 s', '1 m 35 s', '2m 10s'."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().lower()
    total = 0.0
    ok = False
    for m in re.finditer(r"(\d+)\s*m", s):
        total += int(m.group(1)) * 60
        ok = True
    for m in re.finditer(r"(\d+)\s*s", s):
        total += int(m.group(1))
        ok = True
    return total if ok else None


def _parse_number(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _header_map(header: list[str]) -> dict[str, int]:
    """Map normalized header label -> column index."""
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        key = re.sub(r"\s+", " ", str(h or "").strip().lower())
        out[key] = i
    return out


def _col(hmap: dict[str, int], *candidates: str) -> int | None:
    for cand in candidates:
        c = cand.strip().lower()
        if c in hmap:
            return hmap[c]
    for cand in candidates:
        c = cand.strip().lower()
        for k, idx in hmap.items():
            if c in k or k in c:
                return idx
    return None


def _mean(xs: list[float]) -> float | None:
    return sum(xs) / len(xs) if xs else None


def _mode_str(counter: Counter) -> str:
    if not counter:
        return ""
    top = counter.most_common(5)
    return "; ".join(f"{k}:{v}" for k, v in top if k is not None and str(k).strip() != "")


@dataclass
class SessionCalibration:
    sheet: str
    data_origin: str
    n_rows_robinsons_location: int
    n_rows_waltermart_location: int
    n_rows_other_location: int
    n_rows_calibration_subset: int
    demand_t0_s: float
    demand_t1_s: float
    demand_span_raw_s: float
    demand_window_capped: bool
    sim_tmax_s: float
    vol_rb_to_wm: int
    vol_wm_to_rb: int
    mean_dwell_rb_s: float | None
    mean_dwell_wm_s: float | None
    mean_boarding_at_origin_s: float | None
    mean_alighting_at_origin_s: float | None
    mean_arrival_rate_at_origin_s: float | None
    crowding_summary: str
    traffic_condition_summary: str
    bus_companies_summary: str
    routes_summary: str


def load_session_stats(sheet_name: str, *, data_origin: DataOrigin) -> SessionCalibration:
    if not DATA_PATH.is_file():
        raise FileNotFoundError(f"Missing {DATA_PATH}")
    wb = openpyxl.load_workbook(DATA_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Unknown sheet {sheet_name!r}. Have: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    raw_header = [str(h).strip() if h is not None else "" for h in rows[0]]
    hmap = _header_map(raw_header)

    i_time = _col(hmap, "time")
    i_bus = _col(hmap, "bus (company name)", "bus")
    i_route = _col(hmap, "route")
    i_loc = _col(hmap, "location")
    i_board = _col(hmap, "no. of passenger boarding", "boarding")
    i_alight = _col(hmap, "no. of passenger alighting", "alighting")
    i_arrival = _col(hmap, "no. of passenger arrival rate", "arrival rate")
    i_dwell = _col(hmap, "dwell time")
    i_crowd = _col(hmap, "traffic crowding level", "crowding")
    i_traffic = _col(hmap, "traffic condition")

    required = [("Time", i_time), ("Location", i_loc), ("Dwell Time", i_dwell)]
    for label, idx in required:
        if idx is None:
            raise ValueError(f"Sheet {sheet_name!r}: missing column like {label}. Header={raw_header}")

    n_rb = n_wm = n_ot = 0
    drb_all: list[float] = []
    dwm_all: list[float] = []

    for r in rows[1:]:
        if not r:
            continue
        sec = _time_to_seconds(r[i_time])
        if sec is None:
            continue
        loc = _normalize_location(str(r[i_loc] or ""))
        if loc == "robinsons":
            n_rb += 1
        elif loc == "waltermart":
            n_wm += 1
        else:
            n_ot += 1
        d = parse_dwell_seconds(r[i_dwell])
        if d and d > 0:
            if loc == "robinsons":
                drb_all.append(d)
            elif loc == "waltermart":
                dwm_all.append(d)

    subset_times: list[float] = []
    subset_board: list[float] = []
    subset_alight: list[float] = []
    subset_arrival: list[float] = []
    crowd_c = Counter()
    traffic_c = Counter()
    bus_c = Counter()
    route_c = Counter()

    want = "robinsons" if data_origin == "robinsons" else "waltermart"
    n_subset = 0
    for r in rows[1:]:
        if not r:
            continue
        sec = _time_to_seconds(r[i_time])
        if sec is None:
            continue
        loc = _normalize_location(str(r[i_loc] or ""))
        if loc != want:
            continue
        n_subset += 1
        subset_times.append(sec)
        if i_board is not None:
            v = _parse_number(r[i_board])
            if v is not None:
                subset_board.append(v)
        if i_alight is not None:
            v = _parse_number(r[i_alight])
            if v is not None:
                subset_alight.append(v)
        if i_arrival is not None:
            v = _parse_number(r[i_arrival])
            if v is not None:
                subset_arrival.append(v)
        if i_crowd is not None and r[i_crowd] is not None:
            crowd_c[str(r[i_crowd]).strip()] += 1
        if i_traffic is not None and r[i_traffic] is not None:
            traffic_c[str(r[i_traffic]).strip()] += 1
        if i_bus is not None and r[i_bus] is not None:
            bus_c[str(r[i_bus]).strip()[:60]] += 1
        if i_route is not None and r[i_route] is not None:
            route_c[str(r[i_route]).strip()[:80]] += 1

    if not subset_times:
        raise ValueError(
            f"No rows with Location matching {data_origin!r} in {sheet_name!r}. "
            f"Have Robinsons={n_rb}, Waltermart={n_wm} location rows."
        )

    t_min = min(subset_times)
    t_max = max(subset_times)
    obs_rel = [t - t_min for t in subset_times]
    span_raw = max(float(max(obs_rel) - min(obs_rel)), MIN_DEMAND_SPAN_S)
    capped = False
    span_use = span_raw
    if MAX_DEMAND_WINDOW_CAP_S is not None and span_raw > MAX_DEMAND_WINDOW_CAP_S:
        span_use = MAX_DEMAND_WINDOW_CAP_S
        capped = True

    mean_dwell_rb = _mean(drb_all)
    mean_dwell_wm = _mean(dwm_all)

    vol_rb_to_wm = max(n_rb, 1)
    vol_wm_to_rb = max(n_wm, 1)

    origin_label = "robinsons_location" if data_origin == "robinsons" else "waltermart_location"

    return SessionCalibration(
        sheet=sheet_name,
        data_origin=origin_label,
        n_rows_robinsons_location=n_rb,
        n_rows_waltermart_location=n_wm,
        n_rows_other_location=n_ot,
        n_rows_calibration_subset=n_subset,
        demand_t0_s=0.0,
        demand_t1_s=span_use,
        demand_span_raw_s=span_raw,
        demand_window_capped=capped,
        sim_tmax_s=span_use + SIM_CLEARANCE_S,
        vol_rb_to_wm=vol_rb_to_wm,
        vol_wm_to_rb=vol_wm_to_rb,
        mean_dwell_rb_s=mean_dwell_rb,
        mean_dwell_wm_s=mean_dwell_wm,
        mean_boarding_at_origin_s=_mean(subset_board),
        mean_alighting_at_origin_s=_mean(subset_alight),
        mean_arrival_rate_at_origin_s=_mean(subset_arrival),
        crowding_summary=_mode_str(crowd_c),
        traffic_condition_summary=_mode_str(traffic_c),
        bus_companies_summary=_mode_str(bus_c),
        routes_summary=_mode_str(route_c),
    )


def print_calibration(cal: SessionCalibration) -> None:
    print(f"\n--- Calibration: {cal.sheet} | data_origin={cal.data_origin} ---")
    print(
        f"  Location counts (whole sheet): Robinsons={cal.n_rows_robinsons_location}, "
        f"Waltermart={cal.n_rows_waltermart_location}, other={cal.n_rows_other_location}"
    )
    print(f"  Rows in this calibration subset: {cal.n_rows_calibration_subset}")
    if cal.demand_window_capped:
        print(
            f"  Demand window: {cal.demand_t0_s:.0f}-{cal.demand_t1_s:.0f} s "
            f"(capped from raw span {cal.demand_span_raw_s:.0f} s)"
        )
    else:
        print(f"  Demand window: {cal.demand_t0_s:.0f}-{cal.demand_t1_s:.0f} s (raw span {cal.demand_span_raw_s:.0f} s)")
    print(f"  Simulation horizon tmax: {cal.sim_tmax_s:.0f} s")
    print(f"  Volumes: RB->WM={cal.vol_rb_to_wm}, WM->RB={cal.vol_wm_to_rb}")
    if cal.mean_dwell_rb_s:
        print(f"  Mean dwell (Robinsons-location rows' dwell at RB): {cal.mean_dwell_rb_s:.1f} s")
    if cal.mean_dwell_wm_s:
        print(f"  Mean dwell (Waltermart-location rows' dwell at WM): {cal.mean_dwell_wm_s:.1f} s")
    if cal.mean_boarding_at_origin_s is not None:
        print(f"  Mean boarding (subset): {cal.mean_boarding_at_origin_s:.2f}")
    if cal.mean_alighting_at_origin_s is not None:
        print(f"  Mean alighting (subset): {cal.mean_alighting_at_origin_s:.2f}")
    if cal.mean_arrival_rate_at_origin_s is not None:
        print(f"  Mean arrival rate (subset): {cal.mean_arrival_rate_at_origin_s:.2f}")
    if cal.crowding_summary:
        print(f"  Crowding (subset top): {cal.crowding_summary}")
    if cal.traffic_condition_summary:
        print(f"  Traffic condition (subset top): {cal.traffic_condition_summary}")
    if cal.bus_companies_summary:
        print(f"  Bus companies (subset top): {cal.bus_companies_summary}")
    if cal.routes_summary:
        print(f"  Routes (subset top): {cal.routes_summary}")


if __name__ == "__main__":
    for sh in SESSION_SHEETS:
        for origin in ("robinsons", "waltermart"):
            try:
                print_calibration(load_session_stats(sh, data_origin=origin))
            except ValueError as e:
                print(f"{sh} / {origin}: {e}")
